# %% [markdown]
# Cameron Indirect Manual Assignment Tool

# %%
# !pip install pandas selenium webdriver-manager openpyxl

# %%
# Librerías
import os
import pandas as pd
import re
import time
from pathlib import Path
from datetime import datetime
import tempfile
import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import tkinter as tk
from tkinter import ttk, messagebox
import subprocess

# %%
# Detectar ruta de OneDrive SLB o Schlumberger
opciones_busqueda = ["SLB", "Schlumberger"]
opciones_regex = "|".join(opciones_busqueda)
pattern = re.compile(fr'(?i)({opciones_regex})')
onedrive_folder = [p for p in Path.home().iterdir() if p.is_dir() and "OneDrive" in p.name]
onedrive_folder_slb = next((p for p in onedrive_folder if pattern.search(p.name)), None)

if not onedrive_folder_slb:
    raise Exception("No se encontró la carpeta de OneDrive de SLB o Schlumberger en este equipo.")

# Rutas de archivos
ruta_base = onedrive_folder_slb / "Documents" / "Codes" / "Cameron Assignment"
df_path = ruta_base / "DF.xlsx"
ah_path = ruta_base / "Assignation History.csv"
resultado_path =  "./data/Resultado.xlsx"
users_path = ruta_base / "Cameron Users.xlsx"
duplicated_users_path = ruta_base / "Cameron Duplicated Users Management Tool" / "Cameron Duplicated Users Result.xlsx"
rdriver = ruta_base / "msedgedriver.exe"

# Importar archivos
df = pd.read_excel(df_path, engine='openpyxl')
ah = pd.read_csv(ah_path)

# Convertir campos de fecha a datetime
df['Date Submitted'] = pd.to_datetime(df['Date Submitted'], format='%m/%d/%Y %H:%M', errors='coerce')
df['Date Created'] = pd.to_datetime(df['Date Created'], format='%m/%d/%Y %H:%M', errors='coerce')
ah['Created'] = pd.to_datetime(ah['Created'], format='%m/%d/%Y %H:%M', errors='coerce')

# Eliminar en "df" los registros cuyos "ID" estén en "SC Number" de "ah"
df = df[~df['ID'].isin(ah['SC Number'])].copy()

# Crear df2 con los campos requeridos y agregar columna Requester
df2 = pd.DataFrame({
    'ID': df['ID'],
    'PR': '',
    'USER': '',
    'BUYER': df['Assigned To'].fillna(''),
    'PF': '',
    'URGENT': 0,
    'Requester': df['Requester'],
    'Assignment Group': '',
    'Company Code': df['Company Code']
    
})

# Campo URGENT
urgente_patron = re.compile(r'urg|urgent|urgente|urgency', re.IGNORECASE)
df2['URGENT'] = df['Title'].apply(lambda x: 1 if urgente_patron.search(str(x)) else 0)

# Cargar Cameron Users.xlsx si existe, si no, crearlo
if users_path.exists():
    users_df = pd.read_excel(users_path, engine='openpyxl')
    if not set(users_df.columns) >= {"Name", "Alias"}:
        users_df = pd.DataFrame(columns=["Name", "Alias"])
else:
    users_df = pd.DataFrame(columns=["Name", "Alias"])
    users_df.to_excel(users_path, index=False)

# Leer lista de usuarios duplicados por campo 'Requester'
duplicated_requesters = []
if duplicated_users_path.exists():
    duplicated_df = pd.read_excel(duplicated_users_path, engine='openpyxl')
    if 'Requester' in duplicated_df.columns:
        # Normalizar a minúsculas y sin espacios para comparación robusta
        duplicated_requesters = duplicated_df['Requester'].astype(str).str.strip().str.lower().tolist()

# Función para asignar usuario solo si NO es usuario duplicado
def asignar_usuario(row):
    requester = str(row['Requester']).strip().lower()
    if requester in duplicated_requesters:
        # Forzar búsqueda en Ariba para usuarios duplicados
        return ''
    else:
        # Buscar en Users.xlsx para no duplicados
        matches = users_df.loc[users_df['Name'].astype(str).str.strip().str.lower() == requester, 'Alias']
        if not matches.empty:
            return matches.values[0]
        else:
            return ''

df2['USER'] = df2.apply(asignar_usuario, axis=1)

# Revisar si hay usuarios por buscar
usuarios_por_buscar = df2[df2['USER'] == '']

if not usuarios_por_buscar.empty:
    print(f"Se deben buscar {len(usuarios_por_buscar)} usuarios en Ariba. Abriendo navegador...")

    # Separar usuarios duplicados vs no duplicados
    non_duplicated = usuarios_por_buscar[~usuarios_por_buscar['Requester'].astype(str).str.strip().str.lower().isin(duplicated_requesters)]
    duplicated = usuarios_por_buscar[usuarios_por_buscar['Requester'].astype(str).str.strip().str.lower().isin(duplicated_requesters)]
    
    # Configuración robusta del navegador usando EdgeChromiumDriverManager
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument(f"--user-data-dir={os.path.join(os.getenv('TEMP'), 'edge_profile')}")
    temp_profile = tempfile.mkdtemp()
    options.add_argument(f"--user-data-dir={temp_profile}")

    # Crear el servicio y el driver usando EdgeChromiumDriverManager
    service = Service(rdriver)
    driver = webdriver.Edge(service=service, options=options)
    wait = WebDriverWait(driver, 60)

    url = 'https://s1.ariba.com/Sourcing/Main/aw?awh=r&awssk=ODTzCxAIbpbHwlV1&realm=schlumberger'
    driver.get(url)

    # Esperar a que la página principal de Ariba esté completamente cargada
    try:
        wait.until(EC.presence_of_element_located((By.XPATH, "//input[@_pl='ID']")))
        print("Página principal de Ariba cargada correctamente.")
    except Exception as e:
        print(f"Error cargando la página principal de Ariba: {e}")
        driver.quit()
        raise

    # Procesar usuarios NO duplicados (agrupados por requester)
    if not non_duplicated.empty:
        requesters_pendientes = non_duplicated['Requester'].drop_duplicates().tolist()
        for requester in requesters_pendientes:
            ids_con_este_requester = df2[(df2['Requester'] == requester) & (df2['USER'] == '')]['ID'].tolist()
            id_val = str(ids_con_este_requester[0])
            print(f"\nBuscando usuario para Requester: {requester} (Ejemplo ID: {id_val})")

            try:
                # Campo ID
                input_id = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@_pl='ID']")))
                input_id.clear()
                input_id.send_keys(id_val)

                # Lupa azul (botón buscar)
                lupa = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='a-srch-bar-search-icon' and @aria-hidden='true']/ancestor::button[1]")))
                lupa.click()
                print("Lupa clickeada.")

                found = False
                for _ in range(3):  # Solo 3 segundos de espera
                    try:
                        enlaces = driver.find_elements(By.XPATH, "//table[contains(@class,'tableBody')]//a[@bh='HL']")
                        for enlace in enlaces:
                            if enlace.is_displayed() and enlace.text.strip() == id_val:
                                driver.execute_script("arguments[0].scrollIntoView(true);", enlace)
                                enlace.click()
                                print("ID encontrado y clickeado.")
                                found = True
                                break
                        if found:
                            break
                    except Exception:
                        pass
                    time.sleep(1)

                if not found:
                    print(f"ID {id_val} NO encontrado. Cambiando fecha...")

                    # XPATH robusto para el campo de fecha
                    try:
                        fecha_input = wait.until(EC.presence_of_element_located((
                            By.XPATH,
                            "//input[contains(@class, 'w-txt') and contains(@title, 'Enter date')]"
                        )))
                    except Exception:
                        fecha_input = wait.until(EC.presence_of_element_located((
                            By.XPATH,
                            "//input[contains(@type, 'text') and contains(@class, 'w-txt')]"
                        )))

                    fecha_input.clear()
                    fecha_input.send_keys("Thu, 1 May, 2025")
                    time.sleep(1)

                    # Botón Search: XPATH robusto y específico
                    search_btn = wait.until(EC.element_to_be_clickable((
                        By.XPATH,
                        "//button[@title='Run this search' and .//span[normalize-space(text())='Search']]"
                    )))
                    driver.execute_script("arguments[0].scrollIntoView(true);", search_btn)
                    search_btn.click()
                    print("Búsqueda con nueva fecha...")

                    for _ in range(3):  # Solo 3 segundos de espera tras cambiar la fecha
                        try:
                            enlaces = driver.find_elements(By.XPATH, "//table[contains(@class,'tableBody')]//a[@bh='HL']")
                            for enlace in enlaces:
                                if enlace.is_displayed() and enlace.text.strip() == id_val:
                                    driver.execute_script("arguments[0].scrollIntoView(true);", enlace)
                                    enlace.click()
                                    found = True
                                    break
                            if found:
                                break
                        except Exception:
                            pass
                        time.sleep(1)

                    if not found:
                        print(f"ID {id_val} no encontrado tras cambiar fecha. Saltando...")
                        try:
                            logo = wait.until(EC.element_to_be_clickable((By.XPATH, "//img[@alt='Company Logo']")))
                            logo.click()
                            time.sleep(2)
                        except Exception as e:
                            print(f"No se pudo hacer clic en el logo de SLB: {e}")
                        continue

                time.sleep(2)

                solicitante_xpath = "(//a[@bh='HL' and contains(@class, 'hoverLink')])[2]"
                solicitante = wait.until(EC.element_to_be_clickable((By.XPATH, solicitante_xpath)))
                solicitante.click()
                print(f"Solicitante clickeado.")

                email = wait.until(EC.presence_of_element_located((By.XPATH, "//a[starts-with(@href, 'mailto:')]")))
                user = email.text.split('@')[0]
                print(f"USER asignado desde Ariba: {user}")

                # Asignar usuario a todos los IDs con ese requester en df2
                df2.loc[(df2['Requester'] == requester) & (df2['USER'] == ''), 'USER'] = user

                # Actualizar Users.xlsx con Requester y alias
                users_df = users_df[users_df['Name'] != requester]
                users_df = pd.concat([users_df, pd.DataFrame([{"Name": requester, "Alias": user}])], ignore_index=True)

                logo = wait.until(EC.element_to_be_clickable((By.XPATH, "//img[@alt='Company Logo']")))
                logo.click()
                time.sleep(2)

            except Exception as e:
                print(f"Error procesando Requester '{requester}' (ID ejemplo {id_val}): {e}")

    # Procesar usuarios DUPLICADOS (individualmente por ID)
    if not duplicated.empty:
        print("\nProcesando usuarios duplicados (búsqueda individual por ID)...")
        for index, row in duplicated.iterrows():
            id_val = str(row['ID'])
            requester = row['Requester']
            print(f"\nBuscando usuario para ID: {id_val} (Requester: {requester})")

            try:
                # Campo ID
                input_id = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@_pl='ID']")))
                input_id.clear()
                input_id.send_keys(id_val)

                # Lupa azul (botón buscar)
                lupa = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='a-srch-bar-search-icon' and @aria-hidden='true']/ancestor::button[1]")))
                lupa.click()
                print("Lupa clickeada.")

                found = False
                for _ in range(3):
                    try:
                        enlaces = driver.find_elements(By.XPATH, "//table[contains(@class,'tableBody')]//a[@bh='HL']")
                        for enlace in enlaces:
                            if enlace.is_displayed() and enlace.text.strip() == id_val:
                                driver.execute_script("arguments[0].scrollIntoView(true);", enlace)
                                enlace.click()
                                print("ID encontrado y clickeado.")
                                found = True
                                break
                        if found:
                            break
                    except Exception:
                        pass
                    time.sleep(1)

                if not found:
                    print(f"ID {id_val} NO encontrado. Cambiando fecha...")
                    try:
                        fecha_input = wait.until(EC.presence_of_element_located((
                            By.XPATH,
                            "//input[contains(@class, 'w-txt') and contains(@title, 'Enter date')]"
                        )))
                    except Exception:
                        fecha_input = wait.until(EC.presence_of_element_located((
                            By.XPATH,
                            "//input[contains(@type, 'text') and contains(@class, 'w-txt')]"
                        )))

                    fecha_input.clear()
                    fecha_input.send_keys("Thu, 1 May, 2025")
                    time.sleep(1)

                    search_btn = wait.until(EC.element_to_be_clickable((
                        By.XPATH,
                        "//button[@title='Run this search' and .//span[normalize-space(text())='Search']]"
                    )))
                    driver.execute_script("arguments[0].scrollIntoView(true);", search_btn)
                    search_btn.click()
                    print("Búsqueda con nueva fecha...")

                    for _ in range(3):
                        try:
                            enlaces = driver.find_elements(By.XPATH, "//table[contains(@class,'tableBody')]//a[@bh='HL']")
                            for enlace in enlaces:
                                if enlace.is_displayed() and enlace.text.strip() == id_val:
                                    driver.execute_script("arguments[0].scrollIntoView(true);", enlace)
                                    enlace.click()
                                    found = True
                                    break
                            if found:
                                break
                        except Exception:
                            pass
                        time.sleep(1)

                    if not found:
                        print(f"ID {id_val} no encontrado tras cambiar fecha. Saltando...")
                        try:
                            logo = wait.until(EC.element_to_be_clickable((By.XPATH, "//img[@alt='Company Logo']")))
                            logo.click()
                            time.sleep(2)
                        except Exception as e:
                            print(f"No se pudo hacer clic en el logo de SLB: {e}")
                        continue

                time.sleep(2)

                solicitante_xpath = "(//a[@bh='HL' and contains(@class, 'hoverLink')])[2]"
                solicitante = wait.until(EC.element_to_be_clickable((By.XPATH, solicitante_xpath)))
                solicitante.click()
                print(f"Solicitante clickeado.")

                email = wait.until(EC.presence_of_element_located((By.XPATH, "//a[starts-with(@href, 'mailto:')]")))
                user = email.text.split('@')[0]
                print(f"USER asignado desde Ariba: {user}")

                
                df2.loc[df2['ID'] == id_val, 'USER'] = user

                # NO actualizar Users.xlsx para usuarios duplicados

                logo = wait.until(EC.element_to_be_clickable((By.XPATH, "//img[@alt='Company Logo']")))
                logo.click()
                time.sleep(2)

            except Exception as e:
                print(f"Error procesando ID '{id_val}': {e}")


    driver.quit()
    shutil.rmtree(temp_profile, ignore_errors=True)
    
else:
    print("Todos los usuarios fueron encontrados en Users.xlsx (excepto duplicados). No es necesario abrir Ariba para no duplicados.")

# Guardar Cameron Users.xlsx (solo para no duplicados)
users_df.to_excel(users_path, index=False)

# Autoajustar columnas de Cameron Users.xlsx
wb_users = load_workbook(users_path)
for ws in wb_users.worksheets:
    for column_cells in ws.columns:
        max_length = 0
        col = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        ws.column_dimensions[col].width = max_length + 2
wb_users.save(users_path)

# Discriminar entre CAM NAM y CAM LAM

# Definir condiciones
condition_lam = df2['Company Code'].str.contains('AE|AR|BO|BR|CL|CO|CW|EC|GY|MX|PA|PE|SR|SX|SG|TT|UY', na=False)
condition_nam = df2['Company Code'].str.contains('CA|US', na=False)

# Asignar valores usando las condiciones
df2.loc[condition_lam, 'Assignment Group'] = 'CAMERON LAM'
df2.loc[condition_nam, 'Assignment Group'] = 'CAMERON NAM'

# Verificar si quedaron códigos sin asignar
if df2['Assignment Group'].isnull().any():
    print("No se pudo categorizar algunos Company Code dentro de Cameron NAM / LAM.")


# Eliminar columna Requester de df2 antes de exportar
df2 = df2.drop(columns=['Requester'])

# Validar días de creación respecto a la fecha actual
# Crear el campo Days Created
df['Days Created'] = (df['Date Submitted'].dt.normalize() - df['Date Created'].dt.normalize()).dt.days

# Dividir la data en 2 dataframes, uno para IN OTS y otro para OUT OTS
df_in_ots = df[df['Days Created'] <= 7].copy()
df_out_ots = df[df['Days Created'] >= 8].copy()

# Indicar con mensaje si se encontraron solicitudes OUT OF OTS y cuántas
print(f"Solicitudes OUT OF OTS encontradas: {len(df_out_ots)}")

# Exportar la información
def export_out_of_ots(df_out_ots, df2, ruta_base):
    """
    Exportar los PRs OUT OF OTS a 'Cameron Out of OTS PRs.xlsx' con hojas PR LIST y DATA.
    Evita duplicados y crea validación en CATEGORY.
    Retorna True si creó o agregó info y False si no hizo nada.
    """
    out_ots_path = Path(ruta_base) / "Cameron Out of OTS PRs.xlsx"
    hoja_prlist = "PR LIST"
    hoja_data = "DATA"
    hay_nuevos = False

    if df_out_ots.empty:
        print("No hay solicitudes OUT OF OTS para exportar.")
        return False

    # Asegurar que df_out_ots tenga la columna USER desde df2
    if 'USER' not in df_out_ots.columns or df_out_ots['USER'].isnull().all():
        df_out_ots = df_out_ots.merge(df2[['ID','USER']], how='left', on='ID')

    prlist_export = pd.DataFrame({
        'PR': df_out_ots['ID'],
        'USER': df_out_ots['USER'],
        'DAYS CREATED': df_out_ots['Days Created'],
        'CATEGORY': "",
        'BUYER': df_out_ots['Assigned To']
    })

    categorias = [
        "CAM IND LAM", "CAM IND NAM", "D&I", "FES LAM", "FES NAM",
        "FRS/NFRS", "EM DIR EXPENSES", "EM IND", "IND LAM", "IND NAM", "R&R"
    ]
    categoria_formula = '"' + ','.join(categorias) + '"'

    if not out_ots_path.exists():
        # Si el archivo no existe: crear con títulos y validación
        with pd.ExcelWriter(out_ots_path, engine='openpyxl') as writer:
            prlist_export.to_excel(writer, sheet_name=hoja_prlist, index=False)
            df_out_ots.to_excel(writer, sheet_name=hoja_data, index=False)

        wb = load_workbook(out_ots_path)
        ws = wb[hoja_prlist]
        dv = DataValidation(type="list", formula1=categoria_formula, allow_blank=True, showDropDown=True)
        ws.add_data_validation(dv)
        dv.add('C2:C1000')
        wb.save(out_ots_path)
        hay_nuevos = True
        print(f"Archivo '{out_ots_path.name}' creado con {len(df_out_ots)} solicitudes OUT OF OTS.")
    else:
        # Si el archivo existe: poner solo registros nuevos no duplicados
        wb = load_workbook(out_ots_path)
        ws_prlist = wb[hoja_prlist]
        ws_data = wb[hoja_data]

        pr_ids_existentes = set()
        # Identifica IDs existentes en PR LIST, columna A, filas desde 2 hasta max_row
        for row in ws_prlist.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                pr_ids_existentes.add(row[0])

        pr_nuevos = prlist_export[~prlist_export['PR'].isin(pr_ids_existentes)]
        df_out_nuevos = df_out_ots[~df_out_ots['ID'].isin(pr_ids_existentes)]

        if not pr_nuevos.empty:
            hay_nuevos = True
            print(f"Agregando {len(pr_nuevos)} nuevas solicitudes OUT OF OTS al archivo existente.")

            # Añadir nuevas filas a PR LIST
            for _, row in pr_nuevos.iterrows():
                ws_prlist.append(list(row))

            # Añadir nuevas filas a DATA
            for _, row in df_out_nuevos.iterrows():
                ws_data.append(list(row))

            wb.save(out_ots_path)
        else:
            print("No hay nuevas solicitudes OUT OF OTS para agregar al archivo existente.")

    return hay_nuevos

# Exportar OUT OF OTS a "Out of OTS PRs"
nuevos_out_ots = export_out_of_ots(df_out_ots, df2, ruta_base)

# Exportar a Resultado.xlsx (dos hojas: "Reporte" y "Data") solo para IN OTS
df2_in_ots = df2[df2['ID'].isin(df_in_ots['ID'])].copy()
with pd.ExcelWriter(resultado_path, engine='openpyxl', mode='w') as writer:
    df2_in_ots.to_excel(writer, sheet_name='Reporte', index=False)
    df_in_ots.to_excel(writer, sheet_name='Data', index=False)

# Resaltar en rojo las filas con Company Code PA03 o PA02 en la hoja "Reporte"
company_codes_objetivo = [
    "PA03 (Schlumberger SEACO)",
    "PA02 (SLB Overseas, S.A.)"
]
ids_resaltar = df[df['Company Code'].isin(company_codes_objetivo)]['ID'].astype(str).tolist()

wb = load_workbook(resultado_path)
ws = wb['Reporte']

col_id = None
for col in ws.iter_cols(1, ws.max_column):
    if col[0].value == 'ID':
        col_id = col[0].column
        break

if col_id:
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        id_cell = row[col_id - 1]
        if str(id_cell.value) in ids_resaltar:
            for cell in row:
                cell.fill = red_fill

wb.save(resultado_path)

# Autoajustar columnas de las hojas "Reporte" y "Data" en Resultado.xlsx
wb_resultado = load_workbook(resultado_path)
for sheet_name in ["Reporte", "Data"]:
    ws = wb_resultado[sheet_name]
    for column_cells in ws.columns:
        max_length = 0
        col = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        ws.column_dimensions[col].width = max_length + 2
wb_resultado.save(resultado_path)

# Autoajustar columnas de las hojas "PR LIST" y "Data" en Cameron Out of OTS PRs.xlsx
out_ots_path = ruta_base / "Cameron Out of OTS PRs.xlsx"

if out_ots_path.exists():
    wb_out_ots = load_workbook(out_ots_path)
    for sheet_name in ["PR LIST", "DATA"]:
        if sheet_name in wb_out_ots.sheetnames:
            ws = wb_out_ots[sheet_name]
            for column_cells in ws.columns:
                max_length = 0
                col = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        cell_length = len(str(cell.value)) if cell.value is not None else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                ws.column_dimensions[col].width = max_length + 2
    wb_out_ots.save(out_ots_path)

# Confirmar archivos exportados
print("Archivos exportados.")

# Abrir resultados
# os.startfile(str(resultado_path))
if nuevos_out_ots:
    os.startfile(str(ruta_base / "Cameron Out of OTS PRs.xlsx"))

print("¡Proceso finalizado correctamente!")



